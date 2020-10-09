# Importing necessary packages
from openpyxl import *
import openpyxl
from tkinter import *
from tkinter import messagebox as mb
import os.path
import xlrd
import re

# Retrieving and storing the data from excel
curPath = os.getcwd()
print(curPath)
excelFile = curPath + '\list_address.xlsx'
wb = xlrd.open_workbook(excelFile)
sheet = wb.sheet_by_index(0)
no_rows = sheet.nrows
no_columns = sheet.ncols

# Writing the work to the database
workb = load_workbook('list_address.xlsx')
path = 'list_address.xlsx'
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
ws = workb["Sheet1"]
sheetv = workb.active
row = ws.max_row+1


# Phno validation function
def is_valid(ph):
    # 1) Begins with 0 or 91
    # 2) Then contains 7 or 8 or 9.
    # 3) Then contains 9 digits
    pattern = re.compile("(0/91)?[7-9][0-9]{9}")
    return pattern.match(ph)


# Pincode validation function
def is_valid_pincode(pin):
    print(pin)
    regex = "^[1-9]{1}[0-9]{2}\\s{0,1}[0-9]{3}$"
    p = re.compile(regex)
    m = re.match(p, pin)
    print(m)
    if m is None:
        return False
    else:
        return True


# Storing the valid addresses  in the database
def insert_ven_details(market, street, area, city, pin, phno):
    m_row = sheetv.max_row
    for i in range(1, m_row + 1):
        sheet_obj.cell(row=i, column=no_columns)
    # Checking for the duplicate entries by checking the phone number
    dup = 0
    for i in range(1, m_row + 1):
        cell_obj = sheet_obj.cell(row=i+1, column=no_columns)
        if phno == str(cell_obj.value):
            dup = 1
            break
        else:
            # Resetting to zero if there is no duplicate
            dup = 0
    # If there is a duplicate found - Record exist
    if dup == 1:
        print('Record already exits')
        mb.showwarning("Error", "The record already exists")
        ms = mb.askquestion('Exit Application', 'Do you want to exit the application', icon='warning')
        if ms == 'yes':
            root.destroy()
        else:
            # If the user wants to continue delete all the entries
            market_entry.delete(0, END)
            street_entry.delete(0, END)
            area_entry.delete(0, END)
            city_entry.delete(0, END)
            pin_entry.delete(0, END)
            phno_entry.delete(0, END)
    # if there is no duplicate insert each input to excel
    else:
        for i in range(1, ws.max_row+1):
            wcell1 = ws.cell(row, 1)
            wcell1.value = str(market)
            wcell1 = ws.cell(row, 2)
            wcell1.value = str(street)
            wcell1 = ws.cell(row, 3)
            wcell1.value = str(area)
            wcell1 = ws.cell(row, 4)
            wcell1.value = str(city)
            wcell1 = ws.cell(row, 5)
            wcell1.value = str(pin)
            wcell1 = ws.cell(row, 6)
            wcell1.value = str(phno)
            print('Your details are inserted!!')
            mb.showinfo("Thank You", "Your details have been recorded."
                                     "  Thank you for choosing us")
            root.destroy()
            break
    # saving the data to excel sheet
    workb.save('list_address.xlsx')


# A window to prompting for the vendor details
def vendor_details():
    global root
    root = Tk()
    root.title("Testing window")
    root.geometry('600x300+400+150')
    label = Label(root, text="Enter Your Details", font=("calibre", 10, "bold"))
    label.grid(row=0, column=1, columnspan=2, pady=10)
    # Getting all the necessary details
    a1 = Label(root, text="Market Name: ", justify=RIGHT)
    a1.grid(row=1, column=0, sticky=E, pady=5)
    global market_entry
    market_entry = Entry(root, borderwidth=2, width=50, justify=LEFT)
    market_entry.grid(row=1, column=1, columnspan=3, padx=5, pady=5, sticky=W)
    a2 = Label(root, text="Street No and name: ", justify=RIGHT)
    a2.grid(row=3, column=0, sticky=E, pady=5)
    global street_entry
    street_entry = Entry(root, borderwidth=2, width=70, justify=LEFT)
    street_entry.grid(row=3, column=1, columnspan=3, padx=5, pady=5, sticky=W)
    a3 = Label(root, text="Locality & Area: ", justify=RIGHT)
    a3.grid(row=5, column=0, sticky=E, pady=5)
    global area_entry
    area_entry = Entry(root, borderwidth=2, width=70, justify=LEFT)
    area_entry.grid(row=5, column=1, columnspan=3, padx=5, pady=5, sticky=W)
    a4 = Label(root, text="City: ", justify=RIGHT)
    a4.grid(row=7, column=0, sticky=E, pady=5)
    global city_entry
    city_entry = Entry(root, borderwidth=2, width=70, justify=LEFT)
    city_entry.grid(row=7, column=1, columnspan=3, padx=5, pady=5, sticky=W)
    a5 = Label(root, text="Zip Code: ", justify=RIGHT)
    a5.grid(row=9, column=0, sticky=E, pady=5)
    global pin_entry
    pin_entry = Entry(root, borderwidth=2, width=20, justify=LEFT)
    pin_entry.grid(row=9, column=1, columnspan=3, padx=5, pady=5, sticky=W)
    a6 = Label(root, text="Phone Number: ", justify=RIGHT)
    a6.grid(row=11, column=0, sticky=E, pady=5)
    global phno_entry
    phno_entry = Entry(root, borderwidth=2, width=30, justify=LEFT)
    phno_entry.grid(row=11, column=1, columnspan=3, padx=5, pady=5, sticky=W)

    # Checking the all fields or entered and if it is valid
    def submit_check():
        market_name = market_entry.get()
        street_name = street_entry.get()
        area_name = area_entry.get()
        city_name = city_entry.get()
        pin_name = pin_entry.get()
        phno_name = phno_entry.get()
        if len(market_name) != 0 and len(street_name) != 0 and len(area_name) != 0 and len(city_name) != 0 and len(pin_name) != 0 and len(phno_name) != 0:
            if is_valid_pincode(pin_name) and is_valid(phno_name):
                print("Valid")
                insert_ven_details(market_name, street_name, area_name, city_name, pin_name, phno_name)
            elif is_valid_pincode(pin_name) and not is_valid(phno_name):
                mb.showerror("Error", "Enter a valid Phone number")
                phno_entry.delete(0, END)
            elif not is_valid_pincode(pin_name) and is_valid(phno_name):
                mb.showerror("Error", "Enter a valid Pincode")
                pin_entry.delete(0, END)
            else:
                mb.showerror("Error", "Enter a valid Phone Number and Pincode")
                phno_entry.delete(0, END)
                pin_entry.delete(0, END)

        else:
            mb.showwarning("Warning", "Fill all the required fields")
    submit_but = Button(root, text="SUBMIT", command=submit_check, width=12)
    submit_but.grid(row=15, column=1, pady=10)
    root.mainloop()
